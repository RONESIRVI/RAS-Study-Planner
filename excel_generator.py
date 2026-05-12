import os
import glob
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def get_base_paths():
    # Priority: 
    # 1. 'data' and 'output' subfolders
    # 2. Sibling 'राजस्थान का इतिहास' folder
    # 3. Desktop legacy path
    
    out_dir = "output"
    os.makedirs(out_dir, exist_ok=True)
    
    # Search for consolidated files
    search_dirs = [
        "data",
        ".",
        os.path.join("..", "राजस्थान का इतिहास", "Final_PYQ_Output"),
        r'C:\Users\jlpms\OneDrive\Desktop\राजस्थान का इतिहास\Final_PYQ_Output'
    ]
    
    consolidated_dir = "."
    for d in search_dirs:
        if os.path.exists(d):
            files = glob.glob(os.path.join(d, "**", "*Master_Consolidated*.xlsx"), recursive=True)
            if files:
                consolidated_dir = d
                break
                
    return out_dir, consolidated_dir

def generate_topic_excel(topic_name):
    out_dir, consolidated_dir = get_base_paths()
    if not os.path.exists(out_dir): os.makedirs(out_dir)
    
    all_questions = []
    for f in glob.glob(os.path.join(consolidated_dir, "*.xlsx")):
        if "~$" in f or "Master_Consolidated" not in f: continue
        try:
            wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
            ws = wb['PYQ Bank']
            # Smart Match: Clean topic name (remove brackets)
            clean_topic = topic_name.split('(')[0].strip().lower()
            keywords = [k for k in clean_topic.split() if len(k) > 2]
            
            for row in ws.iter_rows(min_row=4, values_only=True):
                target_text = f"{str(row[1])} {str(row[2])}".lower()
                # If the main topic name is in the row, or at least 2 keywords match
                if clean_topic in target_text or (keywords and sum(1 for k in keywords if k in target_text) >= 1):
                    all_questions.append(row)
        except: pass

    if not all_questions: return None

    new_wb = Workbook()
    ws = new_wb.active
    headers = ['Q.No.', 'Topic', 'Sub-Topic', 'Question', '(1) Option A', '(2) Option B', '(3) Option C', '(4) Option D', 'Ans No.', '✅ Correct Answer', '📖 व्याख्या', 'Self Test', 'Revision']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")

    for i, q_row in enumerate(all_questions, 2):
        for j, val in enumerate(q_row[:13], 1):
            cell = ws.cell(row=i, column=j)
            cell.value = val
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            
    safe_name = "".join([c for c in topic_name if c.isalnum() or c==' ']).rstrip()
    excel_path = os.path.join(out_dir, f"PYQ_{safe_name}.xlsx")
    new_wb.save(excel_path)
    return excel_path
