import openpyxl
import os
import glob
import requests
from datetime import datetime, timedelta, date

SHEET_ID = "1Zo81TfPcU09ErH7g-bj-4TksqceuBmiL"
DOWNLOAD_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

def get_tracker_path():
    target_path = os.path.join("data", "Master_Tracker_Live.xlsx")
    os.makedirs("data", exist_ok=True)
    try:
        r = requests.get(DOWNLOAD_URL, timeout=15)
        if r.status_code == 200:
            with open(target_path, "wb") as f:
                f.write(r.content)
            return target_path
    except:
        pass
    return target_path

def get_adaptive_tasks(target_date=None):
    if target_date is None:
        # The planner generates a plan for tomorrow, so look up revisions scheduled for tomorrow
        target_date = (datetime.now() + timedelta(days=1)).date()
    tracker_file = get_tracker_path()
    wb = openpyxl.load_workbook(tracker_file, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    
    # 1. Main Sheet
    main_sheet = next((s for s in ['📋 Master Tracker', 'Master Tracker'] if s in sheet_names), sheet_names[0])
    ws = wb[main_sheet]
    classes_list, revisions = [], []
    for row in ws.iter_rows(min_row=4, max_row=200, values_only=True):
        if not row or len(row) < 14: continue
        if row[1] and str(row[13]).strip().lower() != "done" and len(classes_list) < 2:
            classes_list.append({'subject': str(row[0]), 'topic': str(row[1])})
            revisions.append({'subject': str(row[0]), 'topic': f"{row[1]} (Same Day Rev)"})

    # 2. Smart Revision Search
    rev_sheet_name = next((s for s in sheet_names if 'revision' in s.lower()), None)
    if rev_sheet_name:
        # Avoid printing sheet name directly if it has emojis
        print(f"DEBUG: Found Revision Sheet.")
        rev_ws = wb[rev_sheet_name]
        for row in rev_ws.iter_rows(min_row=4, max_row=200, values_only=True):
            if not row or len(row) < 5: continue
            sub, top = str(row[0]), str(row[1])
            if not top or top == "None": continue
            
            for d_idx in [3, 5, 7, 9, 11]:
                if d_idx >= len(row): continue
                raw_val = row[d_idx]
                parsed_date = None
                if isinstance(raw_val, datetime): parsed_date = raw_val.date()
                elif isinstance(raw_val, str) and "/" in raw_val:
                    for fmt in ["%d/%m/%Y", "%m/%d/%Y"]:
                        try:
                            parsed_date = datetime.strptime(raw_val.strip(), fmt).date()
                            break
                        except: pass
                
                if parsed_date == target_date:
                    done_val = str(row[d_idx+1]).strip().lower() if d_idx+1 < len(row) else ""
                    if done_val != "done":
                        label = f"R{(d_idx-1)//2}" if d_idx < 11 else "Final"
                        revisions.append({'subject': sub, 'topic': f"{top} ({label})"})

    image_data = []
    for idx, c in enumerate(classes_list):
        image_data.append({'sr': idx+1, 'task': f'CLASSES {idx+1}', 'subject': c['subject'], 'topic': c['topic']})
    image_data.append({'sr': len(image_data)+1, 'task': 'REVISION', 'revisions': revisions})
    image_data.append({'sr': len(image_data)+1, 'task': 'Analysis', 'subject': 'PYQ Review', 'topic': 'Previous Year Questions Analysis'})
    
    return image_data, classes_list

def get_weekly_roadmap():
    return ["Next topics are loaded dynamically from Master Tracker."]
