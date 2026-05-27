import os
import glob
import openpyxl
import re
import sys

sys.stdout.reconfigure(encoding='utf-8')

INPUT_DIR = r'R:\Final_PYQ_Output'

def clean_text(text):
    if text is None: return ""
    return str(text).strip()

def split_options(text):
    if not text: return ["", "", "", ""]
    text = str(text)
    options = []
    found_any = False
    
    # Try pattern like (1) Option ... (2) Option ...
    for i in range(1, 5):
        p1 = rf'\({i}\)\s*(.*?)(?=\s*\({i+1}\)|\s*$)'
        p2 = rf'\({chr(64+i)}\)\s*(.*?)(?=\s*\({chr(64+i+1)}\)|\s*$)'
        m1 = re.search(p1, text, re.DOTALL)
        m2 = re.search(p2, text, re.DOTALL)
        if m1:
            options.append(m1.group(1).strip())
            found_any = True
        elif m2:
            options.append(m2.group(1).strip())
            found_any = True
        else:
            options.append("")
            
    # Try pattern like 1. Option ... 2. Option ...
    if not found_any:
        options = []
        for i in range(1, 5):
            p = rf'{i}\.\s*(.*?)(?=\s*{i+1}\.|\s*$)'
            m = re.search(p, text, re.DOTALL)
            if m:
                options.append(m.group(1).strip())
                found_any = True
            else:
                options.append("")
                
    # Try pattern like A. Option ... B. Option ...
    if not found_any:
        options = []
        for i in range(1, 5):
            p = rf'{chr(64+i)}\.\s*(.*?)(?=\s*{chr(64+i+1)}\.|\s*$)'
            m = re.search(p, text, re.DOTALL)
            if m:
                options.append(m.group(1).strip())
                found_any = True
            else:
                options.append("")
                
    return options

def clean_question_text(q_text):
    if not q_text: return ""
    q_text = str(q_text).strip()
    # Strip options listed at the end of the question text
    q_text = re.sub(r'\s*\(\d\)\s*.*$', '', q_text)
    q_text = re.sub(r'\s*\([A-D]\)\s*.*$', '', q_text)
    return q_text.strip()

def parse_file(file_path, default_subtopic):
    print(f"\nParsing: {os.path.basename(file_path)}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        for name in wb.sheetnames:
            if "प्रश्न" in name or "Questions" in name or "Bank" in name:
                sheet = wb[name]
                break
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as e:
        print(f"Error: {e}")
        return []
        
    if not rows: return []
    
    q_col, opt1, opt2, opt3, opt4, opts_merged, ans_col, ans_text_col, exp_col, topic_col, subtopic_col = -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1
    start_row = 0
    
    for idx, r in enumerate(rows[:5]):
        # Skip empty or title rows (title rows only have 1 populated cell)
        non_empty_count = sum(1 for c in r if c is not None and str(c).strip() != "")
        if non_empty_count < 2:
            continue
            
        row_strs = [str(c).lower() if c else "" for c in r]
        for i, s in enumerate(row_strs):
            if "question" in s or "प्रश्न" in s:
                if "option" not in s and "विकल्प" not in s:
                    q_col = i
            elif "विकल्प 1" in s or "option a" in s or "option 1" in s or "विकल्प a" in s or "option\n(a)" in s or "option(a)" in s:
                opt1 = i
            elif "विकल्प 2" in s or "option b" in s or "option 2" in s or "विकल्प b" in s or "option\n(b)" in s or "option(b)" in s:
                opt2 = i
            elif "विकल्प 3" in s or "option c" in s or "option 3" in s or "विकल्प c" in s or "option\n(c)" in s or "option(c)" in s:
                opt3 = i
            elif "विकल्प 4" in s or "option d" in s or "option 4" in s or "विकल्प d" in s or "option\n(d)" in s or "option(d)" in s:
                opt4 = i
            elif "विकल्प" in s or "option" in s:
                opts_merged = i
            elif "सही उत्तर" in s or "correct answer" in s or "answer" in s or "ans" in s:
                if "text" not in s and "पाठ" not in s:
                    ans_col = i
            elif "उत्तर पाठ" in s or "correct answer text" in s:
                ans_text_col = i
            elif "व्याख्या" in s or "explanation" in s or "exp" in s or "definition" in s or "source" in s or "vyakhya" in s:
                exp_col = i
            elif "विषय" in s or "topic" in s:
                topic_col = i
            elif "अध्याय" in s or "sub-topic" in s or "subtopic" in s or "ch #" in s:
                subtopic_col = i
                
        if q_col != -1 and ans_col != -1:
            start_row = idx + 1
            break

    if q_col == -1:
        if len(rows[0]) >= 13:
            q_col, opt1, opt2, opt3, opt4, ans_col, ans_text_col, exp_col, topic_col, subtopic_col = 3, 4, 5, 6, 7, 8, 9, 10, 1, 2
            start_row = 3
        elif len(rows[0]) >= 10:
            q_col, opt1, opt2, opt3, opt4, ans_col, exp_col, topic_col, subtopic_col = 3, 4, 5, 6, 7, 8, 9, 1, 2
            start_row = 1
        elif len(rows[0]) >= 3:
            q_col, ans_col = 1, 2
            if len(rows[0]) >= 4: exp_col = 3
            start_row = 1
        else:
            q_col, ans_col = 0, 1
            start_row = 0

    print(f"  Detected mapping: q_col={q_col}, opts={opt1},{opt2},{opt3},{opt4}, opts_merged={opts_merged}, ans_col={ans_col}, exp={exp_col}, start_row={start_row}")

    parsed_data = []
    for r in rows[start_row:]:
        if not any(r) or len(r) <= max(q_col, ans_col): continue
        
        q_text = clean_text(r[q_col])
        ans_raw = clean_text(r[ans_col])
        if not q_text or not ans_raw: continue
        
        q_text_lower = q_text.lower().strip()
        if q_text_lower in ["", "उत्तर (answer)", "question", "प्रश्न", "question / प्रश्न", "topic", "s.no.", "s.no", "answer", "explanation", "s.no. / क्र.सं.", "topic / विषय", "sub-topic", "व्याख्या (vyakhya)"]:
            continue
            
        topic_ovr = clean_text(r[topic_col]) if topic_col != -1 and topic_col < len(r) else ""
        sub_ovr = clean_text(r[subtopic_col]) if subtopic_col != -1 and subtopic_col < len(r) else ""
        
        opt_list = ["", "", "", ""]
        if opt1 != -1 and opt1 < len(r) and opt2 < len(r) and opt3 < len(r) and opt4 < len(r):
            opt_list = [clean_text(r[opt1]), clean_text(r[opt2]), clean_text(r[opt3]), clean_text(r[opt4])]
        elif opts_merged != -1 and opts_merged < len(r):
            opt_list = split_options(clean_text(r[opts_merged]))
        
        if not any(opt_list) and ("(1)" in q_text or "(A)" in q_text or "1." in q_text):
            opt_list = split_options(q_text)
            q_text = clean_question_text(q_text)
            
        ans_no = ""
        digit_match = re.search(r'\b([1-4])\b', ans_raw)
        char_match = re.search(r'\b([A-D])\b', ans_raw, re.IGNORECASE)
        
        if digit_match:
            ans_no = int(digit_match.group(1))
        elif char_match:
            ans_no = ord(char_match.group(1).upper()) - 64
            
        ans_text = ""
        if ans_no and 0 < ans_no <= 4 and opt_list[ans_no - 1]:
            ans_text = opt_list[ans_no - 1]
        else:
            ans_text = ans_raw
            for i, opt in enumerate(opt_list):
                if opt and (opt.lower() == ans_text.lower() or ans_text.lower() in opt.lower()):
                    ans_no = i + 1
                    ans_text = opt
                    break
        
        if ans_text_col != -1 and ans_text_col < len(r) and r[ans_text_col]:
            ans_text = clean_text(r[ans_text_col])
            
        ans_text = re.sub(r'^\([1-4]\)\s*', '', ans_text)
        ans_text = re.sub(r'^\([A-D]\)\s*', '', ans_text)
        
        exp_text = clean_text(r[exp_col]) if exp_col != -1 and exp_col < len(r) else ""
        
        parsed_data.append({
            'Question': q_text,
            'Options': opt_list,
            'AnsNo': ans_no,
            'Answer': ans_text,
            'Explanation': exp_text
        })
        if len(parsed_data) >= 3:
            break
            
    return parsed_data

# Test on one file from different folders
test_files = [
    r'R:\Final_PYQ_Output\A - राजस्थान का इतिहास\Rajasthan_History_Ch7_Ch8_v2_PYQ_Format.xlsx',
    r'R:\Final_PYQ_Output\B - कला एवं संस्कृति\Rajasthan_KalaSanskriti_82-89_QA_PYQ_Format.xlsx',
    r'R:\Final_PYQ_Output\B - कला एवं संस्कृति\राजस्थान कला एवं संस्कृति 1-9_PYQ_Format.xlsx',
    r'R:\Final_PYQ_Output\B - कला एवं संस्कृति\राजस्थान_सामान्य_ज्ञान_प्रश्नोत्तर.xlsx',
    r'R:\Final_PYQ_Output\COMPUTER\Disha_Computer_Questions.xlsx',
    r'R:\Final_PYQ_Output\हिंदी\Anekarthak_Ch9_Questions.xlsx',
    r'R:\Final_PYQ_Output\राजस्थान राजव्यवस्था\Mukhyamantri_QA.xlsx',
    r'R:\Final_PYQ_Output\📝 English\📝 English_Grammar_PYQ_FINAL_355Q.xlsx'
]

for f in test_files:
    if os.path.exists(f):
        res = parse_file(f, "test")
        print("  First parsed row:")
        if res:
            print("    Q:", repr(res[0]['Question']))
            print("    Opts:", res[0]['Options'])
            print("    AnsNo:", res[0]['AnsNo'], "AnsText:", repr(res[0]['Answer']))
            print("    Exp:", repr(res[0]['Explanation']))
        else:
            print("    Empty result!")
    else:
        print(f"File not found: {f}")
