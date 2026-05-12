import openpyxl
from datetime import datetime

tracker_file = r"f:\AIR-01 RAS\Study_Automation_System\data\Master_Tracker_Live.xlsx"
try:
    wb = openpyxl.load_workbook(tracker_file, data_only=True, read_only=True)
    sheet = wb['Master Tracker']
    
    header_row = 3
    all_headers = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    headers = [str(c).lower().strip() if c else "" for c in all_headers]
    print(f"Headers: {headers}")
    
    count = 0
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row) or len(row) < 2: continue
        section = str(row[0]).strip() if row[0] else ""
        topic = str(row[1]).strip() if row[1] else ""
        Action = str(row[12]).strip().lower() if len(row) > 12 and row[12] else ""
        
        if status != "done":
            print(f"Found Pending: Section={section}, Topic={topic}, action={action}")
            count += 1
            if count > 10: break
except Exception as e:
    print(f"Error: {e}")
