import os
import glob
import openpyxl
import re
import sys
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

INPUT_DIR = r'R:\Final_PYQ_Output'
OUTPUT_DIR = INPUT_DIR
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Styles & Colors
COLOR_DARK_BLUE = "1A237E"
COLOR_LIGHT_BLUE = "E8EAF6"
COLOR_WHITE = "FFFFFF"
COLOR_TEAL = "004D40"
COLOR_PURPLE = "4A148C"
COLOR_ORANGE = "E65100"
COLOR_GREEN = "1B5E20"
COLOR_RED = "B71C1C"
COLOR_LIGHT_GRAY = "FAFAFA"

def get_border(style="thin"):
    return Border(
        left=Side(style=style, color="B0BEC5"),
        right=Side(style=style, color="B0BEC5"),
        top=Side(style=style, color="B0BEC5"),
        bottom=Side(style=style, color="B0BEC5")
    )

def clean_text(text):
    if text is None: return ""
    return str(text).strip()

def split_options(text):
    if not text: return ["", "", "", ""]
    text = str(text)
    options = []
    found_any = False
    
    # Try pattern like (1) Option ... (2) Option ...
    for i in range(1, 5):
        p1 = rf'\({i}\)\s*(.*?)(?=\s*\({i+1}\)|\s*$)'
        p2 = rf'\({chr(64+i)}\)\s*(.*?)(?=\s*\({chr(64+i+1)}\)|\s*$)'
        m1 = re.search(p1, text, re.DOTALL)
        m2 = re.search(p2, text, re.DOTALL)
        if m1:
            options.append(m1.group(1).strip())
            found_any = True
        elif m2:
            options.append(m2.group(1).strip())
            found_any = True
        else:
            options.append("")
            
    # Try pattern like 1. Option ... 2. Option ...
    if not found_any:
        options = []
        for i in range(1, 5):
            p = rf'{i}\.\s*(.*?)(?=\s*{i+1}\.|\s*$)'
            m = re.search(p, text, re.DOTALL)
            if m:
                options.append(m.group(1).strip())
                found_any = True
            else:
                options.append("")
                
    # Try pattern like A. Option ... B. Option ...
    if not found_any:
        options = []
        for i in range(1, 5):
            p = rf'{chr(64+i)}\.\s*(.*?)(?=\s*{chr(64+i+1)}\.|\s*$)'
            m = re.search(p, text, re.DOTALL)
            if m:
                options.append(m.group(1).strip())
                found_any = True
            else:
                options.append("")
                
    return options

def clean_question_text(q_text):
    if not q_text: return ""
    q_text = str(q_text).strip()
    # Strip options listed at the end of the question text
    q_text = re.sub(r'\s*\(\d\)\s*.*$', '', q_text)
    q_text = re.sub(r'\s*\([A-D]\)\s*.*$', '', q_text)
    return q_text.strip()

def find_column_mappings(header_row):
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
    q_col, opt1, opt2, opt3, opt4, opts_merged, ans_col, ans_text_col, exp_col, topic_col, subtopic_col = -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1
    
    # 1. Exact matches or specific terms first
    for i, h in enumerate(headers):
        if h in ["question", "प्रश्न", "question / प्रश्न"]:
            q_col = i
        elif h in ["उत्तर पाठ", "correct answer text", "ans text"]:
            ans_text_col = i
        elif h in ["व्याख्या", "explanation", "exp", "vyakhya"]:
            exp_col = i
        elif h in ["विषय", "topic", "topic / विषय"]:
            topic_col = i
        elif h in ["अध्याय", "sub-topic", "subtopic", "sub-topic / उप-विषय"]:
            subtopic_col = i

    # 2. Options
    for i, h in enumerate(headers):
        if h in ["विकल्प 1", "option a", "option 1", "विकल्प a", "विकल्प (1)", "option (a)", "option\n(a)", "option(a)"]:
            opt1 = i
        elif h in ["विकल्प 2", "option b", "option 2", "विकल्प b", "विकल्प (2)", "option (b)", "option\n(b)", "option(b)"]:
            opt2 = i
        elif h in ["विकल्प 3", "option c", "option 3", "विकल्प c", "विकल्प (3)", "option (c)", "option\n(c)", "option(c)"]:
            opt3 = i
        elif h in ["विकल्प 4", "option d", "option 4", "विकल्प d", "विकल्प (4)", "option (d)", "option\n(d)", "option(d)"]:
            opt4 = i
        elif h in ["विकल्प", "option", "options"]:
            opts_merged = i

    # 3. Correct Answer
    for i, h in enumerate(headers):
        if h in ["सही उत्तर", "correct answer", "answer", "ans", "उत्तर"]:
            if i != ans_text_col:
                ans_col = i

    # 4. Fallbacks
    if q_col == -1:
        for i, h in enumerate(headers):
            if "question" in h or "प्रश्न" in h:
                if not any(x in h for x in ["संख्या", "प्रकार", "id", "option", "विकल्प", "src", "स्रोत"]):
                    q_col = i
                    break
    if ans_col == -1:
        for i, h in enumerate(headers):
            if "सही उत्तर" in h or "correct answer" in h or "answer" in h or "ans" in h or "उत्तर" in h:
                if i != ans_text_col and not any(x in h for x in ["पाठ", "text", "संख्या", "प्रकार"]):
                    ans_col = i
                    break
    if exp_col == -1:
        for i, h in enumerate(headers):
            if "व्याख्या" in h or "explanation" in h or "exp" in h or "vyakhya" in h:
                exp_col = i
                break
    if topic_col == -1:
        for i, h in enumerate(headers):
            if "विषय" in h or "topic" in h:
                topic_col = i
                break
    if subtopic_col == -1:
        for i, h in enumerate(headers):
            if "अध्याय" in h or "subtopic" in h or "sub-topic" in h or "ch #" in h:
                subtopic_col = i
                break
                
    return q_col, opt1, opt2, opt3, opt4, opts_merged, ans_col, ans_text_col, exp_col, topic_col, subtopic_col

def is_header_row(q_text, ans_raw):
    q = q_text.lower().strip()
    a = ans_raw.lower().strip() if ans_raw else ""
    header_terms = [
        "question", "प्रश्न", "s.no", "cr.no", "sr.no", "s.no.", "क्र.सं.", "क.स.", "क्र. सं.", "क्र०सं०",
        "विषय", "topic", "sub-topic", "subtopic", "अध्याय", "ch #", "chapter", "page", "पेज",
        "उत्तर", "answer", "correct answer", "सही उत्तर", "व्याख्या", "explanation", "exp", "vyakhya",
        "difficulty", "difficulty range", "difficulty level", "self test", "revision count", "result", "time"
    ]
    for term in header_terms:
        if q == term or q == f"{term} (question)" or q == f"प्रश्न ({term})" or q == f"{term} / प्रश्न" or q == f"प्रश्न / {term}":
            return True
            
    if re.match(r'^(?:प्रश्न|question|s\.?no|क्र\.?सं)\b', q):
        if any(term in q for term in ["संख्या", "प्रकार", "id", "option", "विकल्प", "src", "स्रोत", "header"]):
            return True
            
    if "प्रश्न" in q and "उत्तर" in a:
        return True
    if "question" in q and "answer" in a:
        return True
        
    return False

def extract_opt1_from_q(q_text, opt_list):
    if not q_text:
        return q_text, opt_list
    if not opt_list[0] and opt_list[1]:
        pattern = r'\s*(?:\(\s*\)|\(\s*(?:1|a|a|0|i|i|\|\||॥|i|:|:\s*:|o)\s*\))\s*([^()]*?)$'
        match = re.search(pattern, q_text, re.IGNORECASE)
        if match:
            extracted_val = match.group(1).strip()
            if extracted_val and len(re.sub(r'[\s_.,\-—:;]+', '', extracted_val)) > 0:
                opt_list[0] = extracted_val
                q_text = q_text[:match.start()].strip()
    return q_text, opt_list

def parse_file(file_path, default_subtopic):
    print(f"  Parsing: {os.path.basename(file_path)}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"  [ERROR] Loading failed: {e}")
        return []
        
    matching_sheets = []
    if len(wb.sheetnames) == 1:
        matching_sheets.append(wb.active)
    else:
        for name in wb.sheetnames:
            if any(k in name for k in ["प्रश्न", "Questions", "Bank", "सभी", "ALL", "प्रश्नोत्तरी"]):
                matching_sheets.append(wb[name])
    if not matching_sheets:
        matching_sheets.append(wb.active)
        
    combined_data = []
    for sheet in matching_sheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows: continue
        
        q_col, opt1, opt2, opt3, opt4, opts_merged, ans_col, ans_text_col, exp_col, topic_col, subtopic_col = -1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1
        start_row = 0
        
        for idx, r in enumerate(rows[:5]):
            non_empty_count = sum(1 for c in r if c is not None and str(c).strip() != "")
            if non_empty_count < 2:
                continue
            
            q_col, opt1, opt2, opt3, opt4, opts_merged, ans_col, ans_text_col, exp_col, topic_col, subtopic_col = find_column_mappings(r)
            if q_col != -1 and ans_col != -1:
                start_row = idx + 1
                break

        if q_col == -1:
            if len(rows[0]) >= 13:
                q_col, opt1, opt2, opt3, opt4, ans_col, ans_text_col, exp_col, topic_col, subtopic_col = 3, 4, 5, 6, 7, 8, 9, 10, 1, 2
                start_row = 3
            elif len(rows[0]) >= 10:
                q_col, opt1, opt2, opt3, opt4, ans_col, exp_col, topic_col, subtopic_col = 3, 4, 5, 6, 7, 8, 9, 1, 2
                start_row = 1
            elif len(rows[0]) >= 3:
                q_col, ans_col = 1, 2
                if len(rows[0]) >= 4: exp_col = 3
                start_row = 1
            else:
                q_col, ans_col = 0, 1
                start_row = 0

        for r in rows[start_row:]:
            if not any(r) or len(r) <= max(q_col, ans_col): continue
            
            q_text = clean_text(r[q_col])
            ans_raw = clean_text(r[ans_col])
            if not q_text or not ans_raw: continue
            
            if is_header_row(q_text, ans_raw):
                continue
                
            topic_ovr = clean_text(r[topic_col]) if topic_col != -1 and topic_col < len(r) else ""
            sub_ovr = clean_text(r[subtopic_col]) if subtopic_col != -1 and subtopic_col < len(r) else ""
            
            opt_list = ["", "", "", ""]
            if opt1 != -1 and opt1 < len(r) and opt2 < len(r) and opt3 < len(r) and opt4 < len(r):
                opt_list = [clean_text(r[opt1]), clean_text(r[opt2]), clean_text(r[opt3]), clean_text(r[opt4])]
            elif opts_merged != -1 and opts_merged < len(r):
                opt_list = split_options(clean_text(r[opts_merged]))
                
            if not any(opt_list) and ("(1)" in q_text or "(A)" in q_text or "1." in q_text):
                opt_list = split_options(q_text)
                q_text = clean_question_text(q_text)
                
            orig_q = q_text
            q_text, opt_list = extract_opt1_from_q(q_text, opt_list)
            if not q_text and orig_q:
                q_text = orig_q
            
            ans_no = ""
            digit_match = re.search(r'\b([1-4])\b', ans_raw)
            char_match = re.search(r'\b([A-D])\b', ans_raw, re.IGNORECASE)
            
            if digit_match:
                ans_no = int(digit_match.group(1))
            elif char_match:
                ans_no = ord(char_match.group(1).upper()) - 64
                
            ans_text = ""
            if ans_no and 0 < ans_no <= 4 and opt_list[ans_no - 1]:
                ans_text = opt_list[ans_no - 1]
            else:
                ans_text = ans_raw
                for i, opt in enumerate(opt_list):
                    if opt and (opt.lower() == ans_text.lower() or ans_text.lower() in opt.lower()):
                        ans_no = i + 1
                        ans_text = opt
                        break
            
            if ans_text_col != -1 and ans_text_col < len(r) and r[ans_text_col]:
                ans_text = clean_text(r[ans_text_col])
                
            cleaned_ans = re.sub(r'^\([1-4]\)\s*', '', ans_text).strip()
            cleaned_ans = re.sub(r'^\([A-D]\)\s*', '', cleaned_ans).strip()
            if cleaned_ans:
                ans_text = cleaned_ans
                
            if not ans_text and ans_raw:
                ans_text = ans_raw
                
            exp_text = clean_text(r[exp_col]) if exp_col != -1 and exp_col < len(r) else ""
            
            combined_data.append({
                'Question': q_text,
                'Options': opt_list,
                'AnsNo': ans_no,
                'Answer': ans_text,
                'Explanation': exp_text,
                'TopicOverride': topic_ovr,
                'SubTopicOverride': sub_ovr
            })
            
    return combined_data

def format_pyq_questions_sheet(ws, data, title_name):
    ws.views.sheetView[0].showGridLines = True
    
    # Row 2: Title
    ws.merge_cells('A2:M2')
    title_cell = ws['A2']
    title_cell.value = f"📝  {title_name} — Previous Year Questions (PYQ) — Complete Answer Bank with Definitions"
    title_cell.font = Font(name="Arial", size=16, bold=True, color=COLOR_WHITE)
    title_cell.fill = PatternFill(start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 40
    
    # Row 3: Headers
    headers = [
        'Q.No.', 'Topic / विषय', 'Sub-Topic', 'Question / प्रश्न', 
        '(1) Option A', '(2) Option B', '(3) Option C', '(4) Option D', 
        'Ans\nNo.', '✅ Correct Answer', '📖 व्याख्या (Definition)', 'Self Test\n✓/✗', 'Revision\nCount'
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill(start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="medium", color="FFFFFF"),
            right=Side(style="medium", color="FFFFFF"),
            top=Side(style="medium", color="FFFFFF"),
            bottom=Side(style="medium", color="FFFFFF")
        )
    ws.row_dimensions[3].height = 30
    
    # Data Rows
    for idx, item in enumerate(data, 4):
        row_data = [
            idx - 3,
            item['Topic'],
            item['SubTopic'],
            item['Question'],
            item['Options'][0],
            item['Options'][1],
            item['Options'][2],
            item['Options'][3],
            item['AnsNo'],
            item['Answer'],
            item['Explanation'],
            "",
            0
        ]
        
        ws.row_dimensions[idx].height = 35
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.value = val
            cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_DARK_BLUE)
            cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")
            cell.border = get_border("thin")
            
            # Alignments
            if col_idx in [4, 11]:  # Question, Explanation
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
    # Column Widths
    widths = {
        "A": 6.0, "B": 22.0, "C": 16.0, "D": 45.0, "E": 22.0, "F": 13.0, 
        "G": 13.0, "H": 13.0, "I": 8.0, "J": 22.0, "K": 50.0, "L": 12.0, "M": 13.0
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def format_topic_index_sheet(ws, data, title_name):
    ws.views.sheetView[0].showGridLines = True
    
    # Row 2: Title
    ws.merge_cells('B2:F2')
    title_cell = ws['B2']
    title_cell.value = "🗂️  Topic-wise Index — Chapter Navigation"
    title_cell.font = Font(name="Arial", size=14, bold=True, color=COLOR_WHITE)
    title_cell.fill = PatternFill(start_color=COLOR_TEAL, end_color=COLOR_TEAL, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 35
    
    # Row 3: Headers (Col B to F)
    headers = ['Topic / विषय', 'Sub-Topic', 'Q.Nos.', 'Count', 'Difficulty Range']
    h_colors = [COLOR_TEAL, COLOR_PURPLE, COLOR_DARK_BLUE, COLOR_GREEN, COLOR_RED]
    
    for i, (h, col_color) in enumerate(zip(headers, h_colors), 2):
        cell = ws.cell(row=3, column=i)
        cell.value = h
        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill(start_color=col_color, end_color=col_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = get_border("thin")
    ws.row_dimensions[3].height = 25
    
    # Grouping Data
    grouped = {}
    for idx, item in enumerate(data, 1):
        t = item['Topic']
        st = item['SubTopic']
        if t not in grouped:
            grouped[t] = {}
        if st not in grouped[t]:
            grouped[t][st] = []
        grouped[t][st].append(str(idx))
        
    row_idx = 4
    for topic, subtopics in grouped.items():
        # Topic header row
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        t_cell = ws.cell(row=row_idx, column=2)
        t_cell.value = f"  📌 {topic}"
        t_cell.font = Font(name="Arial", size=11, bold=True, color=COLOR_WHITE)
        t_cell.fill = PatternFill(start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type="solid")
        t_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 28
        
        # Apply borders to merged cells
        for c in range(2, 7):
            ws.cell(row=row_idx, column=c).border = get_border("thin")
            
        row_idx += 1
        
        # Subtopic rows
        for st, qnos in subtopics.items():
            count = len(qnos)
            diff = "सरल | मध्यम" if count < 5 else "सरल | मध्यम | कठिन"
            
            row_data = [topic, st, ", ".join(qnos), count, diff]
            ws.row_dimensions[row_idx].height = 25
            
            for col_idx, val in enumerate(row_data, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = val
                cell.font = Font(name="Arial", size=10, bold=False, color="000000")
                cell.border = get_border("thin")
                
                if col_idx in [2, 3]:  # Topic, Sub-Topic
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_idx == 4:     # Q.Nos
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:                  # Count, Difficulty Range
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            row_idx += 1
            
    # Widths
    ws.column_dimensions["A"].width = 4.0
    ws.column_dimensions["B"].width = 28.0
    ws.column_dimensions["C"].width = 22.0
    ws.column_dimensions["D"].width = 46.5
    ws.column_dimensions["E"].width = 6.3
    ws.column_dimensions["F"].width = 12.0

def format_selftest_sheet(ws, data, title_name):
    ws.views.sheetView[0].showGridLines = True
    
    # Row 2: Title
    ws.merge_cells('A2:G2')
    title_cell = ws['A2']
    title_cell.value = "✏️  Self-Test Answer Sheet — Practice Mode"
    title_cell.font = Font(name="Arial", size=14, bold=True, color=COLOR_WHITE)
    title_cell.fill = PatternFill(start_color=COLOR_TEAL, end_color=COLOR_TEAL, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 35
    
    # Row 3: Headers
    headers = ['Q.No.', 'Question (Short)', 'Your Answer', 'Correct Answer', 'Result', 'Time (sec)', 'Notes']
    h_colors = [COLOR_DARK_BLUE, COLOR_DARK_BLUE, COLOR_ORANGE, COLOR_GREEN, COLOR_RED, COLOR_PURPLE, COLOR_TEAL]
    
    for i, (h, col_color) in enumerate(zip(headers, h_colors), 1):
        cell = ws.cell(row=3, column=i)
        cell.value = h
        cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill(start_color=col_color, end_color=col_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = get_border("thin")
    ws.row_dimensions[3].height = 25
    
    # Data Rows
    for idx, item in enumerate(data, 4):
        short_q = item['Question'][:100] + "..." if len(item['Question']) > 100 else item['Question']
        formula = f'=IF(C{idx}="","⏳",IF(C{idx}=D{idx},"✅ Correct","❌ Wrong"))'
        
        row_data = [idx - 3, short_q, "", item['Answer'], formula, "", ""]
        ws.row_dimensions[idx].height = 28
        
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col_idx)
            cell.value = val
            cell.font = Font(name="Arial", size=10, bold=False, color="000000")
            cell.fill = PatternFill(start_color=COLOR_LIGHT_GRAY, end_color=COLOR_LIGHT_GRAY, fill_type="solid")
            cell.border = get_border("thin")
            
            if col_idx == 2:  # Short Question
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
    # Widths
    widths = {"A": 6.0, "B": 45.0, "C": 22.0, "D": 16.0, "E": 12.0, "F": 20.0, "G": 20.0}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def format_revision_sheet(ws, data, title_name):
    ws.views.sheetView[0].showGridLines = True
    
    # Row 2: Title
    ws.merge_cells('B2:E2')
    title_cell = ws['B2']
    title_cell.value = "🔖  Quick Revision Cards — Key Points"
    title_cell.font = Font(name="Arial", size=11, bold=True, color=COLOR_WHITE)
    title_cell.fill = PatternFill(start_color=COLOR_PURPLE, end_color=COLOR_PURPLE, fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 35
    
    # Row 3: Headers (Col B to E)
    headers = ['No.', '🔑 Key Point / Question', '💡 Answer / Definition', '📌 Exam Appeared']
    for col_idx, h in enumerate(headers, 2):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h
        cell.font = Font(name="Arial", size=11, bold=True, color=COLOR_WHITE)
        cell.fill = PatternFill(start_color=COLOR_PURPLE, end_color=COLOR_PURPLE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = get_border("thin")
    ws.row_dimensions[3].height = 25
    
    # Data Rows
    for idx, item in enumerate(data, 4):
        row_data = [idx - 3, item['Question'], item['Answer'], ""]
        ws.row_dimensions[idx].height = 30
        
        for col_idx, val in enumerate(row_data, 2):
            cell = ws.cell(row=idx, column=col_idx)
            cell.value = val
            cell.font = Font(name="Arial", size=11, bold=False, color="000000")
            cell.border = get_border("thin")
            
            if col_idx in [3, 4]:  # Question, Answer
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
    # Widths
    ws.column_dimensions["A"].width = 4.0
    ws.column_dimensions["B"].width = 6.0
    ws.column_dimensions["C"].width = 55.0
    ws.column_dimensions["D"].width = 30.0
    ws.column_dimensions["E"].width = 22.0

def consolidate_subject(subject_name, folder_name, default_topic, title_name, filter_func=None):
    print(f"\n[SUBJECT] Consolidating {subject_name}...")
    folder_path = os.path.join(INPUT_DIR, folder_name)
    if not os.path.exists(folder_path):
        print(f"[WARNING] Directory does not exist: {folder_path}")
        return
        
    all_data = []
    
    # Scan Excel files
    excel_files = glob.glob(os.path.join(folder_path, "**", "*.xlsx"), recursive=True)
    for f in excel_files:
        fname = os.path.basename(f)
        if "~$" in fname or "_Master_Consolidated" in fname:
            continue
        # Apply custom filters if any
        if filter_func and not filter_func(f):
            continue
            
        default_sub = fname.replace(".xlsx", "").replace("_QA", "").replace("_PYQ_Format", "").strip()
        parsed = parse_file(f, default_sub)
        
        for item in parsed:
            # Set default Topic and SubTopic if override is not found
            item['Topic'] = item['TopicOverride'] if item['TopicOverride'] else default_topic
            item['SubTopic'] = item['SubTopicOverride'] if item['SubTopicOverride'] else default_sub
            all_data.append(item)
            
    if not all_data:
        print(f"[WARNING] No questions found for {subject_name}")
        return
        
    print(f"Total Questions Parsed: {len(all_data)}")
    
    # Create consolidated workbook
    wb = Workbook()
    ws_pyq = wb.active
    ws_pyq.title = "📝 PYQ Questions"
    
    ws_index = wb.create_sheet("🗂️ Topic Index")
    ws_selftest = wb.create_sheet("✏️ Self-Test")
    ws_revision = wb.create_sheet("🔖 Revision Cards")
    
    # Format and populate worksheets
    format_pyq_questions_sheet(ws_pyq, all_data, title_name)
    format_topic_index_sheet(ws_index, all_data, title_name)
    format_selftest_sheet(ws_selftest, all_data, title_name)
    format_revision_sheet(ws_revision, all_data, title_name)
    
    out_path = os.path.join(OUTPUT_DIR, f"{subject_name}_Master_Consolidated.xlsx")
    wb.save(out_path)
    print(f"[SUCCESS] Consolidated Excel saved: {out_path}")
    return out_path

# Execute consolidation for all 8 subjects

# 1. Rajasthan History
def history_filter(fpath):
    fname = os.path.basename(fpath).lower()
    return "history" in fname or "इतिहास" in fname or "पुरातात्विक" in fname

history_path = consolidate_subject(
    "राजस्थान का इतिहास", 
    "A - राजस्थान  RAS", 
    "राजस्थान इतिहास", 
    "राजस्थान इतिहास",
    filter_func=history_filter
)

# Copy consolidated history file to Rajasthan PYQ.xlsx
if history_path and os.path.exists(history_path):
    dest = os.path.join(OUTPUT_DIR, "Rajasthan PYQ.xlsx")
    shutil.copy2(history_path, dest)
    print(f"[COPY] Copied History master to Rajasthan PYQ.xlsx")

# 2. Rajasthan Art & Culture
art_path = consolidate_subject(
    "Rajasthan Art & Culture", 
    "B - कला एवं संस्कृति", 
    "राजस्थान कला एवं संस्कृति", 
    "राजस्थान कला एवं संस्कृति"
)

# 3. COMPUTER
consolidate_subject(
    "COMPUTER", 
    "COMPUTER", 
    "COMPUTER", 
    "कम्प्यूटर (COMPUTER)"
)

# 4. Science
consolidate_subject(
    "science", 
    "science", 
    "Science", 
    "सामान्य विज्ञान (Science)"
)

# 5. Geography
consolidate_subject(
    "भारत भूगोल", 
    "भारत भूगोल", 
    "भारत भूगोल", 
    "भारत भूगोल"
)

# 6. Polity
consolidate_subject(
    "राजस्थान राजव्यवस्था", 
    "राजस्थान राजव्यवस्था", 
    "राजस्थान राजव्यवस्था", 
    "राजस्थान राजव्यवस्था"
)

# 7. Hindi
consolidate_subject(
    "हिंदी", 
    "हिंदी", 
    "हिंदी", 
    "सामान्य हिंदी"
)

# 8. English
consolidate_subject(
    "📝 English", 
    "📝 English", 
    "English", 
    "General English"
)

print("\n--- ALL SUBJECT CONSOLIDATION COMPLETED SUCCESSFULLY! ---")
