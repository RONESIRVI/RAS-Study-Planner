import os
import glob
import openpyxl
from fpdf import FPDF

BASE_DIR = r'C:\Users\jlpms\OneDrive\Desktop\राजस्थान का इतिहास'
OUTPUT_DIR = os.path.join(BASE_DIR, "Study_Automation")
CONSOLIDATED_DIR = os.path.join(BASE_DIR, "Final_PYQ_Output")
FONT_PATH = r"C:\Windows\Fonts\arial.ttf" # Need a font that supports Devanagari if possible, or fallback

class PYQPDF(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 15)
        self.cell(0, 10, 'RAS Previous Year Questions (PYQ) - Topic Wise', 0, 1, 'C')
        self.ln(5)

def generate_topic_pdf(topic_name):
    # Generating PDF...
    all_questions = []
    
    # Search in both History and Art & Culture consolidated files
    for f in glob.glob(os.path.join(CONSOLIDATED_DIR, "*.xlsx")):
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb['PYQ Bank']
            for row in ws.iter_rows(min_row=4, values_only=True):
                # Col 2 is Topic, Col 3 is Sub-Topic, Col 4 is Question
                row_topic = str(row[1]) if row[1] else ""
                row_sub = str(row[2]) if row[2] else ""
                
                if topic_name.lower() in row_topic.lower() or topic_name.lower() in row_sub.lower():
                    all_questions.append({
                        'q': str(row[3]),
                        'opt': [str(row[4]), str(row[5]), str(row[6]), str(row[7])],
                        'ans': str(row[9]),
                        'exp': str(row[10])
                    })
        except Exception as e:
            pass

    if not all_questions:
        return None

    # Create PDF
    pdf = PYQPDF()
    # fpdf2 requires registering a Unicode font for Hindi
    # Since I might not have a perfect Devanagari ttf path, I'll try to use a standard one or ignore errors
    try:
        pdf.add_font('ArialUni', '', r"C:\Windows\Fonts\arial.ttf") # Arial usually has some support
        pdf.set_font('ArialUni', size=12)
    except:
        pdf.set_font('Arial', size=12)
        
    pdf.add_page()
    pdf.cell(0, 10, f"Topic: {topic_name}", 0, 1, 'L')
    pdf.ln(5)
    
    for i, item in enumerate(all_questions, 1):
        # Clean text for PDF (fpdf1 doesn't like some chars, fpdf2 is better)
        pdf.multi_cell(0, 10, f"Q{i}. {item['q']}")
        pdf.ln(2)
        for j, opt in enumerate(item['opt'], 1):
            pdf.cell(0, 10, f"   ({j}) {opt}", 0, 1)
        pdf.set_text_color(0, 128, 0) # Green for answer
        pdf.cell(0, 10, f"Correct Answer: {item['ans']}", 0, 1)
        pdf.set_text_color(0, 0, 0)
        if item['exp'] and item['exp'] != 'None':
            pdf.multi_cell(0, 10, f"Explanation: {item['exp']}")
        pdf.ln(10)
        
        if pdf.get_y() > 250: pdf.add_page()

    safe_name = "".join([c for c in topic_name if c.isalnum() or c==' ']).rstrip()
    pdf_path = os.path.join(OUTPUT_DIR, f"PYQ_{safe_name}.pdf")
    pdf.output(pdf_path)
    return pdf_path

if __name__ == "__main__":
    generate_topic_pdf("राजस्थान का इतिहास")
