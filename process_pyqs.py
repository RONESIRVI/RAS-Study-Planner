import os
import glob
import openpyxl
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
import shutil

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = r'C:\Users\jlpms\OneDrive\Desktop\राजस्थान का इतिहास'
OUTPUT_DIR = os.path.join(BASE_DIR, "Final_PYQ_Output")

# Style Constants
COLOR_DARK_BLUE = "1A237E"
COLOR_LIGHT_BLUE = "E8EAF6"
COLOR_WHITE = "FFFFFF"

HEADER_FONT = Font(name="Arial", size=10, bold=True, color=COLOR_WHITE)
HEADER_FILL = PatternFill(start_color=COLOR_DARK_BLUE, end_color=COLOR_DARK_BLUE, fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium")
)

BODY_FONT = Font(name="Arial", size=10, bold=True, color=COLOR_DARK_BLUE)
BODY_FILL = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")
BODY_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

COL_WIDTHS = {
    "A": 6, "B": 22, "C": 16, "D": 45, "E": 22, "F": 13, "G": 13, "H": 13,
    "I": 8, "J": 22, "K": 50, "L": 12, "M": 13
}

HEADERS = [
    'Q.No.', 'Topic / विषय', 'Sub-Topic', 'Question / प्रश्न', 
    '(1) Option A', '(2) Option B', '(3) Option C', '(4) Option D', 
    'Ans\nNo.', '✅ Correct Answer', '📖 व्याख्या (Definition)', 'Self Test\n✓/✗', 'Revision\nCount'
]

# Ensure output directory exists
os.makedirs(OUTPUT_DIR, exist_ok=True)

def clean_text(text):
    if text is None: return ""
    return str(text).strip()

def split_options(text):
    if not text: return ["", "", "", ""]
    options = []
    found_any = False
    for i in range(1, 5):
        p1 = rf'\({i}\)\s*(.*?)(?=\s*\({i+1}\)|$)'
        p2 = rf'\({chr(64+i)}\)\s*(.*?)(?=\s*\({chr(64+i+1)}\)|$)'
        m1 = re.search(p1, text, re.DOTALL)
        m2 = re.search(p2, text, re.DOTALL)
        if m1:
            options.append(m1.group(1).strip()); found_any = True
        elif m2:
            options.append(m2.group(1).strip()); found_any = True
        else:
            options.append("")
    if not found_any:
        for i in range(1, 5):
            p = rf'{i}\.\s*(.*?)(?=\s*{i+1}\.|$)'
            m = re.search(p, text, re.DOTALL)
            if m:
                options.append(m.group(1).strip()); found_any = True
            else:
                options.append("")
    return options

def parse_file(file_path):
    print(f"Parsing: {os.path.basename(file_path)}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as e:
        print(f"Error loading {file_path}: {e}")
        return []
    
    if not rows: return []
    
    header_detected = False
    if rows[0] and ("S.No" in str(rows[0][0]) or "Question" in str(rows[0][1])):
        header_detected = True
        
    start_row = 1 if header_detected else 0
    data = []
    
    for r in rows[start_row:]:
        if not any(r): continue
        if len(r) < 2: continue
        
        q_text = clean_text(r[1])
        opt_str = ""
        ans_text = ""
        exp_text = ""
        
        if len(r) >= 5:
            opt_str = clean_text(r[2])
            ans_text = clean_text(r[3])
            exp_text = clean_text(r[4])
        elif len(r) >= 3:
            if "(1)" in q_text or "(A)" in q_text:
                opt_str = q_text
                q_match = re.search(r'^(.*?)\s*(?=\(1\)|\(A\))', q_text, re.DOTALL)
                if q_match: q_text = q_match.group(1).strip()
            ans_text = clean_text(r[2])
            if len(r) >= 4: exp_text = clean_text(r[3])

        opt_list = split_options(opt_str)
        if not opt_str and ("(1)" in q_text or "(A)" in q_text):
             opt_list = split_options(q_text)
             q_match = re.search(r'^(.*?)\s*(?=\(1\)|\(A\))', q_text, re.DOTALL)
             if q_match: q_text = q_match.group(1).strip()

        # Numeric Answer Index
        ans_no = ""
        if ans_text in ["1", "2", "3", "4"]:
            ans_no = int(ans_text)
            idx = ans_no - 1
            if idx >= 0 and idx < 4 and opt_list[idx]:
                ans_text = opt_list[idx]
        elif ans_text in ["A", "B", "C", "D"]:
            ans_no = ord(ans_text.upper()) - 64
            idx = ans_no - 1
            if idx >= 0 and idx < 4 and opt_list[idx]:
                ans_text = opt_list[idx]
        else:
            # Try to match ans_text with options to get index
            for i, opt in enumerate(opt_list):
                if opt and opt.lower() == ans_text.lower():
                    ans_no = i + 1
                    break

        data.append({
            'Question': q_text,
            'Option A': opt_list[0], 'Option B': opt_list[1], 
            'Option C': opt_list[2], 'Option D': opt_list[3],
            'AnsNo': ans_no, 'Answer': ans_text, 'Explanation': exp_text
        })
        
    return data

def apply_styles(ws):
    # Column Widths
    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[col].width = width
    
    # Header Row (Row 3)
    for i, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=i)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER
        
    # Title Row (Row 2)
    ws.merge_cells('A2:M2')
    title_cell = ws['A2']
    title_cell.value = "📝  राजस्थान इतिहास — Previous Year Questions (PYQ) — Complete Answer Bank with Definitions"
    title_cell.font = Font(name="Arial", size=14, bold=True, color=COLOR_DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

def process_all():
    src_folders = glob.glob(os.path.join(BASE_DIR, "*इतिहास*")) + glob.glob(os.path.join(BASE_DIR, "*कला*"))
    for folder in src_folders:
        if "Final_PYQ_Output" in folder: continue
        files = glob.glob(os.path.join(folder, "*.xlsx"))
        for f in files:
            fname = os.path.basename(f)
            if "~$" in fname or "PYQ_Format" in fname: continue
            
            try:
                # Guess Topic/Subtopic
                topic = "राजस्थान का इतिहास" if "इतिहास" in folder else "कला एवं संस्कृति"
                sub_topic = fname.replace(".xlsx", "").replace("_QA", "")
                
                parsed_data = parse_file(f)
                if parsed_data:
                    wb = Workbook()
                    ws = wb.active
                    apply_styles(ws)
                    
                    for i, item in enumerate(parsed_data, 1):
                        row_idx = i + 3
                        row_data = [
                            i, topic, sub_topic, item['Question'],
                            item['Option A'], item['Option B'], item['Option C'], item['Option D'],
                            item['AnsNo'], item['Answer'], item['Explanation'], "", 0
                        ]
                        for col_idx, val in enumerate(row_data, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.value = val
                            cell.font = BODY_FONT
                            cell.fill = BODY_FILL
                            cell.alignment = BODY_ALIGN
                            cell.border = BODY_BORDER
                    
                    category = "A - राजस्थान का इतिहास" if "इतिहास" in folder else "B - कला एवं संस्कृति"
                    cat_dir = os.path.join(OUTPUT_DIR, category)
                    os.makedirs(cat_dir, exist_ok=True)
                    
                    out_path = os.path.join(cat_dir, fname.replace(".xlsx", "_PYQ_Format.xlsx"))
                    wb.save(out_path)
                    print(f"Formatted: {os.path.basename(out_path)}")
            except Exception as e:
                print(f"Error in {fname}: {e}")

process_all()
print("All files formatted with template styling!")
