import openpyxl
import os
import glob
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

base = r'C:\Users\jlpms\OneDrive\Desktop\राजस्थान का इतिहास'
tracker_dirs = glob.glob(os.path.join(base, '*Master Tracker*'))
pyq_path = os.path.join(tracker_dirs[0], 'Rajasthan PYQ.xlsx')
wb = openpyxl.load_workbook(pyq_path)
ws = wb.active

def get_style(cell):
    color_val = None
    if cell.font.color:
        if hasattr(cell.font.color, 'rgb') and isinstance(cell.font.color.rgb, str):
            color_val = cell.font.color.rgb
        else:
            color_val = str(cell.font.color.value)

    fill_val = None
    if cell.fill and hasattr(cell.fill, 'start_color'):
        if hasattr(cell.fill.start_color, 'rgb') and isinstance(cell.fill.start_color.rgb, str):
            fill_val = cell.fill.start_color.rgb
        else:
            fill_val = str(cell.fill.start_color.value)

    return {
        "font_name": cell.font.name,
        "font_size": cell.font.size,
        "bold": cell.font.bold,
        "color": color_val,
        "fill": fill_val,
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrapText
        },
        "border": {
            "left": cell.border.left.style,
            "right": cell.border.right.style,
            "top": cell.border.top.style,
            "bottom": cell.border.bottom.style
        }
    }

print("=== Row 3 Style (A3) ===")
print(json.dumps(get_style(ws['A3']), indent=2))

print("\n=== Row 4 Style (A4) ===")
print(json.dumps(get_style(ws['A4']), indent=2))

print("\n=== Column Widths ===")
widths = {}
for i in range(1, 14):
    letter = openpyxl.utils.get_column_letter(i)
    widths[letter] = ws.column_dimensions[letter].width
print(json.dumps(widths, indent=2))
