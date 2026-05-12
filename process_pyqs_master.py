import os
import glob
import openpyxl
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = r'C:\Users\jlpms\OneDrive\Desktop\राजस्थान का इतिहास'
OUTPUT_DIR = os.path.join(BASE_DIR, "Final_PYQ_Output")

# Colors
COLOR_DARK_BLUE = "1A237E"
COLOR_LIGHT_BLUE = "E8EAF6"
COLOR_WHITE = "FFFFFF"
COLOR_TEAL = "004D40"
COLOR_PURPLE = "4A148C"
COLOR_ORANGE = "E65100"
COLOR_GREEN = "1B5E20"
COLOR_RED = "B71C1C"

def get_border(style="thin"):
    return Border(left=Side(style=style), right=Side(style=style), top=Side(style=style), bottom=Side(style=style))

def apply_header_style(cell, bg_color, font_color=COLOR_WHITE, size=10):
    cell.font = Font(name="Arial", size=size, bold=True, color=font_color)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = get_border("medium")

def clean_text(text):
    if text is None: return ""
    return str(text).strip()

def split_options(text):
    if not text: return ["", "", "", ""]
    options = []
    for i in range(1, 5):
        p = rf'\({i}\)\s*(.*?)(?=\s*\({i+1}\)|\s*$)'
        m = re.search(p, text, re.DOTALL)
        if m: options.append(m.group(1).strip())
        else:
            p = rf'\({chr(64+i)}\)\s*(.*?)(?=\s*\({chr(64+i+1)}\)|\s*$)'
            m = re.search(p, text, re.DOTALL)
            if m: options.append(m.group(1).strip())
            else: options.append("")
    return options

def parse_file(file_path):
    print(f"Parsing: {os.path.basename(file_path)}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        best_sheet = wb.active
        for name in wb.sheetnames:
            if "प्रश्न" in name or "Questions" in name:
                best_sheet = wb[name]
                break
        rows = list(best_sheet.iter_rows(values_only=True))
    except Exception as e:
        print(f"Error: {e}")
        return []
    
    if not rows: return []
    
    # Detect Column Mapping
    q_col, opt_col, ans_col, exp_col, topic_col = -1, -1, -1, -1, -1
    start_row = 0
    
    for idx, r in enumerate(rows[:5]):
        row_strs = [str(c).lower() if c else "" for c in r]
        for i, s in enumerate(row_strs):
            if "question" in s or "प्रश्न" in s: q_col = i
            if "option" in s or "विकल्प" in s: opt_col = i
            if "ans" in s or "उत्तर" in s: ans_col = i
            if "exp" in s or "व्याख्या" in s or "source" in s or "vyakhya" in s: exp_col = i
            if "topic" in s or "विषय" in s: topic_col = i
        
        if q_col != -1 and ans_col != -1:
            start_row = idx + 1
            break
            
    # Fallback for 10-column specific structure
    if q_col == -1 and len(rows[0]) >= 10:
        topic_col, q_col, opt_col_start, ans_col, exp_col = 2, 3, 4, 8, 9
    elif q_col == -1:
        q_col, ans_col, exp_col = 0, 1, 2
    
    data = []
    for r in rows[start_row:]:
        if not any(r) or len(r) <= max(q_col, ans_col): continue
        
        q_text = clean_text(r[q_col])
        topic_ovr = clean_text(r[topic_col]) if topic_col != -1 and topic_col < len(r) else ""
        
        # Options
        opt_list = ["", "", "", ""]
        if opt_col != -1 and opt_col < len(r):
            opt_list = split_options(clean_text(r[opt_col]))
        elif q_col == 3 and len(r) >= 8: # Handle 10-col explicit options
            opt_list = [clean_text(r[4]), clean_text(r[5]), clean_text(r[6]), clean_text(r[7])]
            
        ans_text = clean_text(r[ans_col]) if ans_col < len(r) else ""
        exp_text = clean_text(r[exp_col]) if exp_col != -1 and exp_col < len(r) else ""
        
        if not any(opt_list) and ("(1)" in q_text or "(A)" in q_text):
            opt_list = split_options(q_text)
            q_match = re.search(r'^(.*?)\s*(?=\(1\)|\(A\))', q_text, re.DOTALL)
            if q_match: q_text = q_match.group(1).strip()
            
        ans_no = ""
        if ans_text in ["1", "2", "3", "4", "A", "B", "C", "D"]:
            if ans_text in ["1", "2", "3", "4"]: ans_no = int(ans_text)
            else: ans_no = ord(ans_text.upper()) - 64
            if 0 < ans_no <= 4 and opt_list[ans_no-1]: ans_text = opt_list[ans_no-1]
        else:
            for i, o in enumerate(opt_list):
                if o and (o.lower() == ans_text.lower() or ans_text.lower() in o.lower()):
                    ans_no = i + 1
                    break
                    
        data.append({'Question': q_text, 'Options': opt_list, 'AnsNo': ans_no, 'Answer': ans_text, 'Explanation': exp_text, 'TopicOverride': topic_ovr})
    return data

def create_pyq_sheet(wb, data, type_name):
    ws = wb.create_sheet("PYQ Bank")
    ws.merge_cells('A2:M2')
    ws['A2'] = f"📝  {type_name} — Previous Year Questions (PYQ) — Complete Answer Bank"
    ws['A2'].font = Font(name="Arial", size=14, bold=True, color=COLOR_DARK_BLUE)
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    headers = ['Q.No.', 'Topic', 'Sub-Topic', 'Question', '(1) Option A', '(2) Option B', '(3) Option C', '(4) Option D', 'Ans No.', '✅ Correct Answer', '📖 व्याख्या', 'Self Test', 'Revision']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i)
        cell.value = h
        apply_header_style(cell, COLOR_DARK_BLUE)
    for i, item in enumerate(data, 1):
        # Use TopicOverride as Sub-Topic if available
        sub_t = item['TopicOverride'] if item['TopicOverride'] else item['SubTopic']
        row = [i, type_name, sub_t, item['Question']] + item['Options'] + [item['AnsNo'], item['Answer'], item['Explanation'], "", 0]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=i+3, column=c_idx)
            cell.value = val
            cell.font = Font(name="Arial", size=10, bold=True, color=COLOR_DARK_BLUE)
            cell.fill = PatternFill(start_color=COLOR_LIGHT_BLUE, end_color=COLOR_LIGHT_BLUE, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = get_border("thin")
    widths = [6, 20, 20, 45, 20, 20, 20, 20, 8, 20, 50, 10, 10]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

def create_index_sheet(wb, data, topic_name):
    ws = wb.create_sheet("Topic-wise Index", 0)
    ws.merge_cells('A1:E1')
    ws['A1'] = "📂 Topic-wise Index — Chapter Navigation"
    apply_header_style(ws['A1'], COLOR_DARK_BLUE, size=14)
    headers = ['Topic / विषय', 'Sub-Topic', 'Q.Nos.', 'Count', 'Difficulty']
    h_colors = [COLOR_TEAL, COLOR_PURPLE, COLOR_DARK_BLUE, COLOR_GREEN, COLOR_RED]
    for i, (h, col) in enumerate(zip(headers, h_colors), 1):
        cell = ws.cell(row=2, column=i)
        cell.value = h
        apply_header_style(cell, col)
    grouped = {}
    for i, item in enumerate(data, 1):
        sub_t = item['TopicOverride'] if item['TopicOverride'] else item['SubTopic']
        if sub_t not in grouped: grouped[sub_t] = []
        grouped[sub_t].append(str(i))
    row_idx = 3
    ws.merge_cells(f'A{row_idx}:E{row_idx}')
    ws[f'A{row_idx}'] = f"📌 {topic_name}"
    apply_header_style(ws[f'A{row_idx}'], COLOR_DARK_BLUE, size=11)
    row_idx += 1
    for st, qnos in grouped.items():
        row_data = [topic_name, st, ", ".join(qnos), len(qnos), "सरल | मध्यम"]
        for i, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=i)
            cell.value = val
            cell.font = Font(name="Arial", size=10)
            cell.border = get_border()
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1
    widths = [25, 25, 40, 10, 15]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

def create_selftest_sheet(wb, data):
    ws = wb.create_sheet("Self-Test")
    ws.merge_cells('A1:G1')
    ws['A1'] = "🖊️ Self-Test Answer Sheet — Practice Mode"
    apply_header_style(ws['A1'], COLOR_TEAL, size=14)
    headers = ['Q.No.', 'Question (Short)', 'Your Answer', 'Correct Answer', 'Result', 'Time (sec)', 'Notes']
    h_colors = [COLOR_DARK_BLUE, COLOR_DARK_BLUE, COLOR_ORANGE, COLOR_GREEN, COLOR_RED, COLOR_PURPLE, COLOR_TEAL]
    for i, (h, col) in enumerate(zip(headers, h_colors), 1):
        cell = ws.cell(row=2, column=i)
        cell.value = h
        apply_header_style(cell, col)
    for i, item in enumerate(data, 1):
        short_q = item['Question'][:100] + "..." if len(item['Question']) > 100 else item['Question']
        row = [i, short_q, "", item['Answer'], "⏳", "", ""]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=i+2, column=c_idx)
            cell.value = val
            cell.border = get_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions['B'].width = 50

def create_revision_sheet(wb, data):
    ws = wb.create_sheet("Quick Revision")
    ws.merge_cells('A1:D1')
    ws['A1'] = "⚡ Quick Revision Cards — Key Points"
    apply_header_style(ws['A1'], COLOR_PURPLE, size=14)
    headers = ['No.', '🔑 Key Point / Question', '💡 Answer / Definition', '📌 Exam Appeared']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i)
        cell.value = h
        apply_header_style(cell, COLOR_PURPLE)
    for i, item in enumerate(data, 1):
        row = [i, item['Question'], item['Answer'], ""]
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=i+2, column=c_idx)
            cell.value = val
            cell.border = get_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50

def process_consolidated(type_name, folder_name):
    print(f"\nProcessing Folder: {folder_name}")
    folder_path = os.path.join(BASE_DIR, folder_name)
    if not os.path.exists(folder_path): return
    all_data = []
    for f in glob.glob(os.path.join(folder_path, "*.xlsx")):
        if "~$" in f: continue
        parsed = parse_file(f)
        sub_topic = os.path.basename(f).replace(".xlsx", "").replace("_QA", "").replace("_PYQ_Format", "")
        for item in parsed:
            item['SubTopic'] = sub_topic
            all_data.append(item)
    if all_data:
        wb = Workbook(); del wb['Sheet']
        create_index_sheet(wb, all_data, type_name)
        create_pyq_sheet(wb, all_data, type_name)
        create_selftest_sheet(wb, all_data)
        create_revision_sheet(wb, all_data)
        out_path = os.path.join(OUTPUT_DIR, f"{type_name}_Master_Consolidated.xlsx")
        wb.save(out_path)
        print(f"Saved: {os.path.basename(out_path)} ({len(all_data)} Questions)")

process_consolidated("Rajasthan History", "इतिहास")
process_consolidated("Rajasthan Art & Culture", "कला एवं संस्कृति")
print("\nSuccess!")
